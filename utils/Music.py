import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from db import SessionContext
from db.model.Music import Music
from exception import AnalyzeError


class MusicDat:
    ALL_MUSIC = None

    def __init__(self):
        print("MusicDat init...")
        self.wb = openpyxl.load_workbook("utils/djmax.xlsx")
        self.need_header = ["곡명", "BPM", "아티스트"]

    def get_sheets(self):
        return list(map(lambda x: x.upper(), self.wb.sheetnames))

    def get_all_music(self):
        if MusicDat.ALL_MUSIC is None:
            print("all music init")
            with SessionContext() as session:
                all_music = session.query(Music).all()
                MusicDat.ALL_MUSIC = list(map(lambda x: x.music_name, all_music))
        return MusicDat.ALL_MUSIC

    def init_db(self):
        # DB에 데이터를 넣을 때에만 사용
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            dlc_name = sheetname.upper()
            # 첫번째 행은 헤더데이터로 따로 가져옴
            header = {i: h.value for i, h in enumerate(sheet[1]) if h.value in self.need_header}
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                current_music = {v: row[k] for k, v in header.items()}
                current_music["DLC"] = dlc_name
                with SessionContext() as session:
                    music = Music(music_dlc=current_music["DLC"], music_name=current_music["곡명"],
                                  music_artist=current_music["아티스트"], music_bpm=current_music["BPM"])
                    session.add(music)
                    session.commit()


def find_most_similar(text_list, target_text):
    # 텍스트 리스트에 타겟 텍스트를 추가하여 벡터화를 수행합니다.
    texts = text_list + [target_text]

    # TF-IDF 벡터라이저를 사용하여 텍스트를 벡터로 변환합니다.
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(texts)

    # 코사인 유사도를 계산합니다. 마지막 행이 타겟 텍스트의 벡터입니다.
    similarity_matrix = cosine_similarity(tfidf_matrix[-1], tfidf_matrix[:-1])

    # 가장 유사한 텍스트의 인덱스를 찾습니다.
    most_similar_index = similarity_matrix.argmax()

    return text_list[most_similar_index], similarity_matrix[0, most_similar_index]


def most_similar_title(title: str) -> tuple:
    """실제 DB에 있는 음악 제목 중 가장 유사한 제목을 찾아 반환합니다."""
    if title is None:
        raise AnalyzeError("제목이 잘못 표기되었습니다.")
    music_title_list = MusicDat().get_all_music()
    similar_title, similarity_score = find_most_similar(music_title_list, title)
    return similar_title, similarity_score


def get_music_level():
    return {"NORMAL": 10, "HARD": 20, "MAXIMUM": 30, "SC": 40}