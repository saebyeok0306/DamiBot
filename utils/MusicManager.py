import asyncio
from typing import List

import discord
import openpyxl
from discord import Message, Reaction, User, Interaction
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from app import DamiBot
from db import SessionContext
from db.model.Music import Music
from exception import AnalyzeError


class MusicDat:
    ALL_MUSIC = None

    def __init__(self):
        self.wb = openpyxl.load_workbook("utils/djmax.xlsx")
        self.need_header = ["곡명", "BPM", "아티스트"]

    def get_sheets(self):
        return list(map(lambda x: x.upper(), self.wb.sheetnames))

    def get_all_music(self):
        if MusicDat.ALL_MUSIC is None:
            print("all music init")
            with SessionContext() as session:
                all_music = session.query(Music).all()
                MusicDat.ALL_MUSIC = list(map(lambda x: x.music_name, all_music))
        return MusicDat.ALL_MUSIC

    def init_db(self):
        # DB에 데이터를 넣을 때에만 사용
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            dlc_name = sheetname.upper()
            # 첫번째 행은 헤더데이터로 따로 가져옴
            header = {i: h.value for i, h in enumerate(sheet[1]) if h.value in self.need_header}
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                current_music = {v: row[k] for k, v in header.items()}
                current_music["DLC"] = dlc_name
                with SessionContext() as session:
                    music = Music(music_dlc=current_music["DLC"], music_name=current_music["곡명"],
                                  music_artist=current_music["아티스트"], music_bpm=current_music["BPM"])
                    session.add(music)
                    session.commit()


def find_most_similar(text_list, target_text):
    # 텍스트 리스트에 타겟 텍스트를 추가하여 벡터화를 수행합니다.
    texts = text_list + [target_text]

    # TF-IDF 벡터라이저를 사용하여 텍스트를 벡터로 변환합니다.
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(texts)

    # 코사인 유사도를 계산합니다. 마지막 행이 타겟 텍스트의 벡터입니다.
    similarity_matrix = cosine_similarity(tfidf_matrix[-1], tfidf_matrix[:-1])

    # 가장 유사한 텍스트의 인덱스를 찾습니다.
    most_similar_index = similarity_matrix.argmax()

    # FIX:
    # A Lie, Lie Lie를 둘다 1.0으로 나옴.

    return text_list[most_similar_index], similarity_matrix[0, most_similar_index]


def contains_text(text_list, search_text):
    return search_text in text_list


def most_similar_title(title: str) -> tuple:
    """실제 DB에 있는 음악 제목 중 가장 유사한 제목을 찾아 반환합니다."""
    if title is None:
        raise AnalyzeError("제목이 잘못 표기되었습니다.")
    music_title_list = MusicDat().get_all_music()
    for music_title in music_title_list:
        if title == music_title:
            return music_title, 1.0
    similar_title, similarity_score = find_most_similar(music_title_list, title)
    return similar_title, similarity_score


def is_same_title(title: str) -> bool:
    if title in ["Urban Night", "Voyage", "Showdown", "Alone"]:
        return True
    return False


def unify_music_button(button: str):
    button = button.lower()
    match button:
        case "4b" | "4":
            return 4
        case "5b" | "5":
            return 5
        case "6b" | "6":
            return 6
        case "8b" | "8":
            return 8
    return button


def unify_music_level(level: str):
    level = level.lower()
    match level:
        case "normal" | "nm" | "n":
            return "NORMAL"
        case "hard" | "hd" | "h":
            return "HARD"
        case "maximum" | "mx" | "m":
            return "MAXIMUM"
        case "sc" | "s":
            return "SC"
    return level


def get_music_level():
    return {"NORMAL": 10, "HARD": 20, "MAXIMUM": 30, "SC": 40}


def get_music_level_index(level_text: str):
    # lv가 "NORMAL", "HARD", "MAXIMUM", "SC" 중 하나인 경우
    for lv, idx in get_music_level().items():
        if level_text.strip().upper() == lv:
            return idx

    return False


async def same_title_confirm(bot: DamiBot, message: Message, same_title: str, same_title_list: List[Music]) -> Music:
    sel_emoji = ["1️⃣", "2️⃣", "3️⃣", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
    embed = discord.Embed(title=f"{same_title}의 DLC 선택 안내 ",
                          description="DJMAX 내에 같은 이름의 곡이 있어서, 어떤 DLC인지 선택이 필요합니다.",
                          color=0x8d76bc)
    embed.set_thumbnail(url=message.author.display_avatar)
    embed.set_footer(text=f"DJMAX RESPECT V｜기록 관리봇 담이",
                     icon_url=bot.user.display_avatar)

    for idx, title_data in enumerate(same_title_list):
        embed.add_field(name=f"{idx + 1}번",
                        value=f"{title_data.music_artist}\n{title_data.music_dlc}")

    select_message = await message.reply(embed=embed, mention_author=True)
    select_len = len(same_title_list)

    for i in range(select_len):
        await select_message.add_reaction(sel_emoji[i])

    try:
        def check(emoji: Reaction, reaction_user: User):
            return str(emoji) in sel_emoji and reaction_user == message.author and emoji.message.id == select_message.id
        reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

        for i in range(select_len):
            if str(reaction) == sel_emoji[i]:
                await select_message.delete()
                return same_title_list[i]

    except asyncio.TimeoutError:
        await select_message.delete()
        raise TimeoutError("시간이 초과되었습니다.")

    except Exception as e:
        await select_message.delete()
        print("무슨 오류? ", e)


async def same_title_confirm_action(bot: DamiBot, action: Interaction[DamiBot], same_title: str, same_title_list: List[Music]) -> Music:
    sel_emoji = ["1️⃣", "2️⃣", "3️⃣", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
    embed = discord.Embed(title=f"{same_title}의 DLC 선택 안내 ",
                          description="DJMAX 내에 같은 이름의 곡이 있어서, 어떤 DLC인지 선택이 필요합니다.",
                          color=0x8d76bc)
    embed.set_thumbnail(url=action.user.display_avatar)
    embed.set_footer(text=f"DJMAX RESPECT V｜기록 관리봇 담이",
                     icon_url=bot.user.display_avatar)

    for idx, title_data in enumerate(same_title_list):
        embed.add_field(name=f"{idx + 1}번",
                        value=f"{title_data.music_artist}\n{title_data.music_dlc}")

    select_message = await action.followup.send(embed=embed, ephemeral=False)
    select_len = len(same_title_list)

    for i in range(select_len):
        await select_message.add_reaction(sel_emoji[i])

    try:
        def check(emoji: Reaction, reaction_user: User):
            return str(emoji) in sel_emoji and reaction_user == action.user and emoji.message.id == select_message.id
        reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

        for i in range(select_len):
            if str(reaction) == sel_emoji[i]:
                await select_message.delete()
                return same_title_list[i]

    except asyncio.TimeoutError:
        await select_message.delete()
        raise TimeoutError("시간이 초과되었습니다.")

    except Exception as e:
        await select_message.delete()
        print("무슨 오류? ", e)


async def get_music_from_title(bot: DamiBot, message: Message, title_name: str, action: Interaction[DamiBot] = None) -> Music:
    is_same = False
    if is_same_title(title_name):
        is_same = True
    with SessionContext() as session:
        if is_same is True:
            same_title_list = session.query(Music).filter(Music.music_name == title_name).all()
            if action is None:
                return await same_title_confirm(bot, message, title_name, same_title_list)
            else:
                return await same_title_confirm_action(bot, action, title_name, same_title_list)
        else:
            return session.query(Music).filter(Music.music_name == title_name).first()
